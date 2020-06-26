import openpyxl
import os

os.chdir('D:\Documents\Hobby\Automate the Boring Stuff With Python\Excel_Files_To_Edit')


wb = openpyxl.Workbook()    #capital W in Workbook is for creating a new spreadsheet.
wb
#<openpyxl.workbook.workbook.Workbook object at 0x000001631D760BB0>
wb.get_sheet_names()
#['Sheet']
sheet = wb.get_sheet_by_name('Sheet')
sheet
#<Worksheet "Sheet">
sheet['A1'].value
sheet['A1'].value == None
#True
sheet['A1'] = 42
sheet['A2'] = 'Hello'
wb.save('example2.xlsx')    #save a new excel spreadsheet

sheet2 = wb.create_sheet()
wb.get_sheet_names()
#['Sheet', 'Sheet1']
sheet2.title
#'Sheet1'
sheet2.title = 'My New Sheet Name'   #change sheet name
wb.get_sheet_names()
#['Sheet', 'My New Sheet Name']
wb.save('example3.xlsx')

wb.create_sheet(index=0, title = 'My Other Sheet')  #index sets the order for new sheet (starts at 0; title sets the name of the sheet.
#<Worksheet "My Other Sheet">
wb.save('example4.xlsx')
