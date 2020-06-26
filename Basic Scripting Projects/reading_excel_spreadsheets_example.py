import os
import openpyxl
#^install using pip.exe

os.chdir('D:\Documents\Hobby\Automate the Boring Stuff With Python\Excel_Files_To_Edit')
os.path.exists('D:\Documents\Hobby\Automate the Boring Stuff With Python\Excel_Files_To_Edit\example.xlsx')
#True

workbook = openpyxl.load_workbook('example.xlsx')
type(workbook)
#<class 'openpyxl.workbook.workbook.Workbook'>

sheet = workbook.get_sheet_by_name('Sheet1')
type(sheet)
#<class 'openpyxl.worksheet.worksheet.Worksheet'>

workbook.get_sheet_names()
#['Sheet1', 'Sheet2', 'Sheet3']

sheet['A1']
#<Cell 'Sheet1'.A1>
cell = sheet['A1']
cell.value
#datetime.datetime(2015, 4, 5, 13, 34, 2)      #the data type depends on the formatting of the cell
str(cell.value)
#'2015-04-05 13:34:02'
str(sheet['A1'].value)   #similar method
#'2015-04-05 13:34:02'
sheet['B1'].value
#'Apples'
sheet['C1'].value
#73
str(sheet['C1'].value)
#'73'

sheet.cell(row=1,column=2)   #useful for for loops; starts at 1, not 0
#<Cell 'Sheet1'.B1>
sheet['B1']
#<Cell 'Sheet1'.B1>
for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)
"""
1 Apples
2 Cherries
3 Pears
4 Oranges
5 Apples
6 Bananas
7 Strawberries
"""
